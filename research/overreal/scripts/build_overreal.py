#!/usr/bin/env python3
"""Build the standardized OverReal dataset from the raw Drive dump.

Input : data/OverReal/{Attribution,Figurative,Perspectival}/
        each family has Index_<Family>.xlsx (sheet "Prompts & Targets":
        Folder, Prompt, Target, Original, _, Comment) and one subfolder
        per item named <Annotator>_<n>_check containing labeled images.

Image filename convention: <label1>__<label2>.<ext>
  label1: first annotator's label (disruptive/silent/integrated/withheld,
          possibly with numeric suffixes or typos -- normalized here)
  label2: second annotator's verdict: "ok" (agree), "other" (exclude),
          or an override label (disagreement; override wins).
  Files with no "__" were confirmed to be forgotten "ok" marks.
  Special case: "disruptive_silent.png" is a typo for "disruptive__silent".

Output: data/overreal_v1/
  images/<family>/prompt_0001/image_000.<ext>   (family-scoped numeric item ids,
                                        assigned in Index row order)
  metadata.jsonl                        one record per image
  audit_log.csv                         everything that needed a judgment call
  README.md
"""

import csv
import json
import re
import shutil
import sys
from pathlib import Path

from openpyxl import load_workbook

RAW = Path(__file__).resolve().parents[3] / "data" / "OverReal"
OUT = Path(__file__).resolve().parents[3] / "data" / "overreal_v1"

FAMILIES = ["Attribution", "Figurative", "Perspectival"]
CORE_LABELS = {"disruptive", "silent", "integrated", "withheld"}
TYPO_MAP = {
    "disrupted": "disruptive",
    "disprutive": "disruptive",
    "disrputive": "disruptive",
    "disrtuptive": "disruptive",
    "disruptuve": "disruptive",
    "siilent": "silent",
    "intergrated": "integrated",
    "wthheld": "withheld",
}
IMG_EXTS = {".png", ".jpg", ".jpeg"}
# single-underscore filenames confirmed as typos for double underscore
SPECIAL_STEMS = {"disruptive_silent": ("disruptive", "silent")}


def norm_label(raw):
    """Normalize a raw label token: lowercase, strip numeric suffixes, fix typos."""
    s = raw.strip().lower().lstrip("_")      # silent___ok -> second part "_ok"
    s = re.sub(r"\(\d+\)$", "", s)           # ok(1), ok(2): dedup suffix from naming
    s = re.sub(r"[_\s]*\d+$", "", s)         # disruptive_2, silent1, disruptive_01 ...
    s = TYPO_MAP.get(s, s)
    return s


def canon_key(folder):
    """Case/zero-padding-insensitive key: Felix_01 == Felix_1_check == felix_1."""
    name = folder.removesuffix("_check")
    m = re.fullmatch(r"([A-Za-z]+)_0*(\d+)", name)
    return f"{m.group(1).lower()}_{int(m.group(2))}" if m else name.lower()


def parse_stem(stem):
    """Return (label1_raw, label2_raw or None). None means missing 2nd mark."""
    if stem in SPECIAL_STEMS:
        return SPECIAL_STEMS[stem]
    if "__" in stem:
        first, _, second = stem.partition("__")
        return first, second
    return stem, None


def read_index(fam):
    """Folder -> {prompt, target, original, row} from Index_<fam>.xlsx."""
    path = RAW / fam / f"Index_{fam}.xlsx"
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Prompts & Targets"] if "Prompts & Targets" in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = [str(c).strip() if c else "" for c in next(rows)]
    col = {name: i for i, name in enumerate(header)}
    out = {}
    for r, row in enumerate(rows, start=2):
        folder = row[col["Folder"]]
        if folder is None or not str(folder).strip():
            continue
        folder = str(folder).strip()
        get = lambda name: (str(row[col[name]]).strip() if col.get(name) is not None
                            and row[col[name]] is not None else None)
        out[canon_key(folder)] = {
            "prompt": get("Prompt"),
            "target": get("Target"),
            "original": get("Original"),
            "row": r,
        }
    wb.close()
    return out


def main():
    if OUT.exists():
        shutil.rmtree(OUT)
    (OUT / "images").mkdir(parents=True)

    records, anomalies = [], []

    def flag(kind, family, folder, file, detail):
        anomalies.append({"type": kind, "family": family, "folder": folder,
                          "file": file, "detail": detail})

    for fam in FAMILIES:
        fam_lc = fam.lower()
        index = read_index(fam)

        disk = {}  # canon_key -> dir path
        for d in sorted((RAW / fam).iterdir()):
            if not d.is_dir():
                continue
            key = canon_key(d.name)
            if not d.name.endswith("_check"):
                flag("folder_not_checked", fam_lc, d.name, "", "no _check suffix")
            if key in disk:
                flag("duplicate_folder", fam_lc, d.name, "",
                     f"canonical name collides with {disk[key].name}")
                continue
            disk[key] = d

        # item ids follow Index row order; disk-only folders appended after
        ordered = [f for f in index if f in disk]
        for f in index:
            if f not in disk:
                flag("index_row_without_folder", fam_lc, f, "", f"index row {index[f]['row']}")
        for f in sorted(set(disk) - set(index)):
            flag("folder_without_index_row", fam_lc, disk[f].name, "", "kept, no prompt/target")
            ordered.append(f)

        for iid, key in enumerate(ordered, start=1):
            src_dir = disk[key]
            meta = index.get(key, {})
            item_id = f"{fam_lc}/prompt_{iid:04d}"
            annotator = src_dir.name.split("_")[0].capitalize()
            dst_dir = OUT / "images" / fam_lc / f"prompt_{iid:04d}"
            dst_dir.mkdir(parents=True)

            files = sorted(p for p in src_dir.iterdir()
                           if p.suffix.lower() in IMG_EXTS)
            for p in src_dir.iterdir():
                if p.is_file() and p.suffix.lower() not in IMG_EXTS:
                    flag("non_image_file", fam_lc, src_dir.name, p.name, "skipped")

            # pass 1: parse labels
            parsed = []
            for p in files:
                l1_raw, l2_raw = parse_stem(p.stem)
                label_1 = norm_label(l1_raw)
                if label_1 not in CORE_LABELS:
                    flag("bad_label_1", fam_lc, src_dir.name, p.name,
                         f"unrecognized '{l1_raw}', record excluded")
                    label_1 = None

                if l2_raw is None:
                    flag("missing_second_mark", fam_lc, src_dir.name, p.name,
                         "treated as ok (confirmed by author)")
                    verdict = "ok"
                else:
                    verdict = norm_label(l2_raw)

                included, agreement = True, None
                if verdict == "ok":
                    label_2, agreement = label_1, True
                elif verdict == "other":
                    label_2, included = None, False
                elif verdict in CORE_LABELS:
                    label_2, agreement = verdict, False
                else:
                    flag("bad_label_2", fam_lc, src_dir.name, p.name,
                         f"unrecognized '{l2_raw}', record excluded")
                    label_2, included = None, False
                if label_1 is None:
                    included = False
                parsed.append((p, l1_raw, label_1, label_2, agreement, included))

            # resolve is_original: Index "Original" names the Gemini original's
            # filename. Rules (confirmed by author):
            #   1. a single-image folder: that image is the original
            #   2. exact raw-name match preferred ("silent" != "silent_2"),
            #      normalized-label match as fallback (absorbs typos)
            #   3. multiple candidates: lowest-numbered file wins
            #      (disruptive_1 before disruptive_2, __ok before __ok(1))
            original_ref = meta.get("original")
            original_idx = None
            if len(parsed) == 1:
                original_idx = 0
            elif original_ref is not None:
                ref = original_ref.strip().lower()
                hits = [i for i, (_, raw, *_ ) in enumerate(parsed)
                        if raw.strip().lower() == ref]
                if not hits:
                    hits = [i for i, (_, _, l1, *_ ) in enumerate(parsed)
                            if l1 == norm_label(original_ref)]
                if not hits:
                    flag("original_match", fam_lc, src_dir.name, "",
                         f"Original='{original_ref}' matched 0 files")
                else:
                    def num_key(i):
                        stem = parsed[i][0].stem
                        first = re.search(r"(\d+)\s*$", parsed[i][1])
                        paren = re.search(r"\((\d+)\)", stem)
                        return (int(first.group(1)) if first else 0,
                                int(paren.group(1)) if paren else 0, stem)
                    original_idx = min(hits, key=num_key)
                    if len(hits) > 1:
                        flag("original_tiebreak", fam_lc, src_dir.name,
                             parsed[original_idx][0].name,
                             f"{len(hits)} candidates, lowest number chosen")

            # pass 2: copy + emit records
            for seq, (p, _, label_1, label_2, agreement, included) in enumerate(parsed):
                ext = ".jpg" if p.suffix.lower() == ".jpeg" else p.suffix.lower()
                rel = f"images/{fam_lc}/prompt_{iid:04d}/image_{seq:03d}{ext}"
                shutil.copy2(p, OUT / rel)

                records.append({
                    "file_name": rel,
                    "image_id": f"{item_id}/image_{seq:03d}",
                    "family": fam_lc,
                    "item_id": item_id,
                    "source_folder": src_dir.name,
                    "annotator": annotator,
                    "prompt": meta.get("prompt"),
                    "target": meta.get("target"),
                    "label_1": label_1,
                    "label_2": label_2,
                    "agreement": agreement,
                    "included": included,
                    "is_original": seq == original_idx,
                    "raw_path": str(p.relative_to(RAW.parent)),
                })

    with open(OUT / "metadata.jsonl", "w", encoding="utf-8") as f:
        for r in records:
            f.write(json.dumps(r, ensure_ascii=False) + "\n")

    with open(OUT / "audit_log.csv", "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["type", "family", "folder", "file", "detail"])
        w.writeheader()
        w.writerows(anomalies)

    n_inc = sum(r["included"] for r in records)
    n_agree = sum(1 for r in records if r["agreement"] is True)
    n_dis = sum(1 for r in records if r["agreement"] is False)
    (OUT / "README.md").write_text(f"""# OverReal v1

Built by `research/overreal/scripts/build_overreal.py` from `data/OverReal/` (raw Drive dump).
Load with: `datasets.load_dataset("imagefolder", data_dir=".")`.

- families: attribution / figurative / perspectival (Others excluded)
- {len(records)} images, {n_inc} included ({len(records) - n_inc} excluded via "other" or bad labels)
- second-annotator verdicts: {n_agree} agree, {n_dis} overridden
- labels: disruptive / silent / integrated / withheld
- `label_1` first annotator, `label_2` adjudicated final label (override wins),
  `agreement` whether annotator 2 said ok, `included=False` -> drop from analysis,
  `is_original` file referenced by the Index "Original" column (Gemini original)
- see `audit_log.csv` for every judgment call made during conversion
""", encoding="utf-8")

    print(f"records: {len(records)}  included: {n_inc}  "
          f"agree: {n_agree}  overridden: {n_dis}  anomalies: {len(anomalies)}")


if __name__ == "__main__":
    sys.exit(main())
